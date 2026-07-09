#!/usr/bin/env python3
"""
gen_coverage_rpt.py - Extract coverage data from *.vdb databases and generate Excel report.
Usage: python3 gen_coverage_rpt.py -m yyy
"""

import argparse
import subprocess
import os
import re
import logging
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# Output directory
# ---------------------------------------------------------------------------
OUTPUT_DIR = "output"
MODULE_DIR = os.path.join(OUTPUT_DIR, "module")
SPLIT_DIR = os.path.join(OUTPUT_DIR, "module_split")

# ---------------------------------------------------------------------------
# Logging setup
# ---------------------------------------------------------------------------
os.makedirs(MODULE_DIR, exist_ok=True)
os.makedirs(SPLIT_DIR, exist_ok=True)
logging.basicConfig(
    filename=os.path.join(OUTPUT_DIR, 'trace.log'),
    level=logging.INFO,
    format='%(asctime)s - %(message)s',
    datefmt='%Y-%m-%d %H:%M:%S'
)


def log(msg: str) -> None:
    """Write message to both console and trace.log."""
    print(msg)
    logging.info(msg)


# ---------------------------------------------------------------------------
# Step 1: run urg
# ---------------------------------------------------------------------------
def run_urg() -> bool:
    """Run urg -dir *.vdb -format text to generate urgReport directory."""
    log("=" * 60)
    log("Step 1: Running urg to generate coverage report")
    log("=" * 60)

    vdb_files = [f for f in os.listdir('.') if f.endswith('.vdb')]
    if not vdb_files:
        log("ERROR: No .vdb files found in current directory")
        return False

    log(f"  Found .vdb file(s): {', '.join(vdb_files)}")

    cmd = "urg -dir *.vdb -format text"
    log(f"  Command: {cmd}")

    result = subprocess.run(cmd, shell=True, stdout=subprocess.PIPE,
                            stderr=subprocess.PIPE, universal_newlines=True)

    if result.returncode != 0:
        log(f"  ERROR: urg failed with return code {result.returncode}")
        if result.stderr:
            log(f"  stderr: {result.stderr}")
        if result.stdout:
            log(f"  stdout: {result.stdout}")
        return False

    log("  urg completed successfully")
    return True


# ---------------------------------------------------------------------------
# Step 2: parse modinfo.txt and generate per-module .txt files
# ---------------------------------------------------------------------------
def parse_modinfo() -> dict:
    """
    Parse urgReport/modinfo.txt and split it into per-module sections.

    Each module section is identified by the pattern:
        ===============================================================================
        Module : nnn
        ===============================================================================

    Returns a dict mapping module_name -> section_text.
    """
    modinfo_path = os.path.join("urgReport", "modinfo.txt")
    if not os.path.exists(modinfo_path):
        log(f"ERROR: {modinfo_path} not found")
        return {}

    log("=" * 60)
    log("Step 2: Parsing modinfo.txt and generating per-module .txt files")
    log("=" * 60)

    with open(modinfo_path, 'r', encoding='utf-8', errors='replace') as fh:
        content = fh.read()

    lines = content.split('\n')

    # Locate every module header:
    #   ===============================================================================
    #   Module : <name>        (name must be non-empty)
    #   ===============================================================================
    SEP = '=' * 79
    module_headers = []          # list of (line_index, module_name)

    i = 0
    while i < len(lines):
        if (lines[i] == SEP and
            i + 2 < len(lines) and
            lines[i + 1].startswith('Module : ') and
            len(lines[i + 1].strip()) > len('Module : ') and   # non-empty name
            lines[i + 2] == SEP):
            mod_name = lines[i + 1].strip()[len('Module : '):]
            module_headers.append((i, mod_name))
            i += 3
        else:
            i += 1

    log(f"  Found {len(module_headers)} module header(s) in modinfo.txt")

    # Extract each module's content (from its header to the next header, or EOF)
    modules = {}
    for idx, (start_line, mod_name) in enumerate(module_headers):
        end_line = (module_headers[idx + 1][0]
                    if idx + 1 < len(module_headers)
                    else len(lines))
        section_text = '\n'.join(lines[start_line:end_line])

        modules[mod_name] = section_text

        # Write per-module .txt file
        out_filename = os.path.join(MODULE_DIR, f"{mod_name}.txt")
        with open(out_filename, 'w', encoding='utf-8') as fh:
            fh.write(section_text)
        log(f"  Saved {out_filename}  ({len(section_text)} chars)")

    return modules


# ---------------------------------------------------------------------------
# Helper: parse SCORE / LINE / COND / TOGGLE / FSM / ASSERT from section
# ---------------------------------------------------------------------------
def parse_scores(section: str) -> dict:
    """
    Extract the coverage scores from a module section.

    Looks for the SCORE header line and reads the values on the following line.
    Format:
        SCORE  LINE   COND   TOGGLE FSM    ASSERT
         11.91  32.59  15.03   0.00   0.00 --
    """
    keys = ['SCORE', 'LINE', 'COND', 'TOGGLE', 'FSM', 'ASSERT']
    scores = dict.fromkeys(keys, '--')

    lines = section.split('\n')
    for i, line in enumerate(lines):
        stripped = line.strip()
        if stripped.startswith('SCORE') and 'LINE' in stripped and i + 1 < len(lines):
            values = lines[i + 1].strip().split()
            for j, key in enumerate(keys):
                if j < len(values):
                    scores[key] = values[j]
            break

    return scores


# ---------------------------------------------------------------------------
# Parsing helpers for Excel generation (Step 4)
# ---------------------------------------------------------------------------

def _parse_module_summary(module_path: str):
    """
    Parse the module .txt file to extract:
      - hierarchy_name  (from "Module self-instances" or "Module Instance" header)
      - subtree_scores  (LINE / TOGGLE / FSM / COND  as decimals from "Instance's subtree")
    Returns (hierarchy_name, subtree_scores_dict).
    """
    with open(module_path, 'r', encoding='utf-8', errors='replace') as fh:
        content = fh.read()

    lines = content.split('\n')
    hierarchy_name = ''
    scores = {'LINE': '--', 'TOGGLE': '--', 'FSM': '--', 'COND': '--'}

    # --- hierarchy name: try "Module self-instances" first ---
    for i, line in enumerate(lines):
        if line.strip() == 'Module self-instances :':
            for j in range(i + 1, min(i + 12, len(lines))):
                if 'SCORE' in lines[j] and 'LINE' in lines[j] and j + 1 < len(lines):
                    parts = lines[j + 1].strip().split()
                    if len(parts) >= 7:
                        hierarchy_name = parts[-1]
                    break
            break

    # fallback: "Module Instance : xxx" header
    if not hierarchy_name:
        for i, line in enumerate(lines):
            if (line.startswith('=' * 79) and i + 1 < len(lines) and
                    lines[i + 1].startswith('Module Instance :')):
                hierarchy_name = lines[i + 1].strip()[len('Module Instance : '):]
                break

    # fallback: module name itself
    if not hierarchy_name:
        for i, line in enumerate(lines):
            if (line.startswith('=' * 79) and i + 1 < len(lines) and
                    lines[i + 1].startswith('Module : ') and
                    len(lines[i + 1].strip()) > len('Module : ')):
                hierarchy_name = lines[i + 1].strip()[len('Module : '):]
                break

    # --- Instance's subtree scores ---
    key_map = ['SCORE', 'LINE', 'COND', 'TOGGLE', 'FSM', 'ASSERT']
    in_subtree = False
    for i, line in enumerate(lines):
        if line.strip() == "Instance's subtree :":
            in_subtree = True
            continue
        if in_subtree and 'SCORE' in line and 'LINE' in line and i + 1 < len(lines):
            values = lines[i + 1].strip().split()
            for j, key in enumerate(key_map):
                if j < len(values) and values[j] != '--':
                    try:
                        scores[key] = float(values[j]) / 100.0
                    except ValueError:
                        scores[key] = values[j]
            break

    # If no subtree scores found, use the module's own first SCORE line
    if all(v == '--' for v in scores.values()):
        for i, line in enumerate(lines):
            if line.strip().startswith('SCORE') and 'LINE' in line and i + 1 < len(lines):
                values = lines[i + 1].strip().split()
                for j, key in enumerate(key_map):
                    if j < len(values) and values[j] != '--':
                        try:
                            scores[key] = float(values[j]) / 100.0
                        except ValueError:
                            scores[key] = values[j]
                break

    return hierarchy_name, scores


def _parse_line_uncovered(line_path: str):
    """
    Parse <module>_line.txt — return list of (line_no: int, code: str)
    for lines marked with '==>' (uncovered).
    """
    items = []
    if not os.path.exists(line_path):
        return items

    with open(line_path, 'r', encoding='utf-8', errors='replace') as fh:
        for line in fh:
            if '==>' not in line:
                continue
            # Format: "71         0/1     ==>                  tx_state <= START;"
            stripped = line.rstrip('\n')
            # Split on '==>'
            parts = stripped.split('==>', 1)
            left = parts[0].strip()
            code = parts[1].strip() if len(parts) > 1 else ''

            # Extract line number from left part (first token)
            left_parts = left.split()
            if left_parts:
                try:
                    line_no = int(left_parts[0])
                    items.append((line_no, code))
                except ValueError:
                    pass
    return items


def _parse_toggle_uncovered(toggle_path: str):
    """
    Parse <module>_toggle.txt — return list of (type_: str, name: str)
    for ports whose Toggle column is 'No' (uncovered).
    """
    items = []
    if not os.path.exists(toggle_path):
        return items

    with open(toggle_path, 'r', encoding='utf-8', errors='replace') as fh:
        content = fh.read()

    # Locate "Port Details" section
    lines = content.split('\n')
    in_port_details = False
    for line in lines:
        if line.strip() == 'Port Details':
            in_port_details = True
            continue
        if not in_port_details:
            continue
        # Stop at "Signal Details" or blank line after the section
        if line.strip() == 'Signal Details':
            break
        if not line.strip():
            continue
        # Skip header line
        if 'Toggle' in line and 'Direction' in line:
            continue
        # Parse: PCLK               No     No          No          INPUT
        parts = line.split()
        if len(parts) >= 5:
            name = parts[0]
            toggle_val = parts[1]       # 'No' or 'Yes'
            direction = parts[-1]       # 'INPUT', 'OUTPUT', 'INOUT'
            if toggle_val == 'No':
                items.append((direction, name))
    return items


def _parse_cond_expressions(cond_path: str):
    """
    Parse <module>_cond.txt — return list of (line_no: int, block_text: str)
    for every EXPRESSION / SUB-EXPRESSION block.
    """
    items = []
    if not os.path.exists(cond_path):
        return items

    with open(cond_path, 'r', encoding='utf-8', errors='replace') as fh:
        content = fh.read()

    lines = content.split('\n')

    # Find every block starting with " LINE       <num>" followed by
    # " EXPRESSION ..." or " SUB-EXPRESSION ..."
    block_starts = []
    for i, line in enumerate(lines):
        stripped = line.strip()
        if stripped.startswith('LINE') and i + 1 < len(lines):
            m = re.match(r'^LINE\s+(\d+)', stripped)
            if m and ('EXPRESSION' in lines[i + 1] or 'SUB-EXPRESSION' in lines[i + 1]):
                block_starts.append((i, int(m.group(1))))

    for idx, (start, line_no) in enumerate(block_starts):
        end = block_starts[idx + 1][0] if idx + 1 < len(block_starts) else len(lines)
        block_text = '\n'.join(lines[start:end]).rstrip()
        items.append((line_no, block_text))

    return items


def _parse_fsm_uncovered(fsm_path: str):
    """
    Parse <module>_fsm.txt — return list of (line_no: int, transition: str)
    for transitions with 'Not Covered' status.
    """
    items = []
    if not os.path.exists(fsm_path):
        return items

    with open(fsm_path, 'r', encoding='utf-8', errors='replace') as fh:
        content = fh.read()

    # Locate the transitions block
    lines = content.split('\n')
    in_transitions = False
    for line in lines:
        if line.strip().startswith('transitions') and 'Line No.' in line:
            in_transitions = True
            continue
        if not in_transitions:
            continue
        if not line.strip():
            break  # blank line ends the transitions block
        if 'Line No.' in line and 'Covered' in line:
            continue  # skip sub-header

        # Format: BIT0->BIT1    94       Not Covered
        parts = line.split()
        if len(parts) >= 3 and parts[-1] == 'Covered':
            transition = parts[0]
            try:
                line_no = int(parts[1])
                items.append((line_no, transition))
            except ValueError:
                pass

    return items


# ---------------------------------------------------------------------------
# Step 4: generate Excel report (matching coverage_template.xlsx)
# ---------------------------------------------------------------------------
def generate_excel_from_template(target_module: str) -> None:
    """Generate Excel report that follows the coverage_template.xlsx layout."""
    log("=" * 60)
    log("Step 4: Generating Excel report from template")
    log("=" * 60)

    module_path = os.path.join(MODULE_DIR, f"{target_module}.txt")
    line_path = os.path.join(SPLIT_DIR, f"{target_module}_line.txt")
    toggle_path = os.path.join(SPLIT_DIR, f"{target_module}_toggle.txt")
    cond_path = os.path.join(SPLIT_DIR, f"{target_module}_cond.txt")
    fsm_path = os.path.join(SPLIT_DIR, f"{target_module}_fsm.txt")

    # --- Parse all data ---
    hierarchy_name, subtree = _parse_module_summary(module_path)
    line_items = _parse_line_uncovered(line_path)
    toggle_items = _parse_toggle_uncovered(toggle_path)
    cond_items = _parse_cond_expressions(cond_path)
    fsm_items = _parse_fsm_uncovered(fsm_path)

    log(f"  Hierarchy     : {hierarchy_name}")
    log(f"  Line uncovered: {len(line_items)}")
    log(f"  Toggle uncov. : {len(toggle_items)}")
    log(f"  Cond blocks   : {len(cond_items)}")
    log(f"  FSM uncov.    : {len(fsm_items)}")

    # --- Build workbook ---
    wb = Workbook()
    ws = wb.active
    ws.title = "Coverage Report"

    # -- Styles --
    section_fill = PatternFill(start_color='D9E2F3', end_color='D9E2F3', fill_type='solid')
    section_font = Font(name='Calibri', bold=True, size=11)
    header_fill = PatternFill(start_color='B4C6E7', end_color='B4C6E7', fill_type='solid')
    header_font = Font(name='Calibri', bold=True, size=10)
    normal_font = Font(name='Calibri', size=10)
    wrap_align = Alignment(horizontal='left', vertical='top', wrap_text=True)
    center_align = Alignment(horizontal='center', vertical='center')
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin'),
    )
    thick_bottom = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='medium'),
    )

    col_widths = [18, 55, 18, 12, 22]

    def _apply_border(ws_row, border=thin_border):
        for c in range(1, 6):
            ws.cell(row=ws_row, column=c).border = border

    def _write_section_header(ws_row, title):
        """Write a merged section header row (e.g. 'Line Coverage')."""
        ws.merge_cells(start_row=ws_row, start_column=1, end_row=ws_row, end_column=5)
        cell = ws.cell(row=ws_row, column=1, value=title)
        cell.font = section_font
        cell.fill = section_fill
        cell.alignment = center_align
        _apply_border(ws_row)
        return ws_row + 1

    def _write_sub_header(ws_row, col_titles):
        """Write the sub-header row (Line | Code | Comment | Approve | Supplementary Materials)."""
        for c, title in enumerate(col_titles, 1):
            cell = ws.cell(row=ws_row, column=c, value=title)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = thin_border
        return ws_row + 1

    # ---- Row 1: title (module identifier) ----
    row = 1
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
    title_cell = ws.cell(row=row, column=1, value=target_module)
    title_cell.font = Font(name='Calibri', bold=True, size=14)
    title_cell.alignment = center_align
    _apply_border(row)

    # ---- Row 2: summary header ----
    row = 2
    sum_headers = ['hierarchy', 'Line', 'Toggle', 'FSM', 'Condition']
    for c, h in enumerate(sum_headers, 1):
        cell = ws.cell(row=row, column=c, value=h)
        cell.font = Font(name='Calibri', bold=True, size=11)
        cell.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        cell.font = Font(name='Calibri', bold=True, size=11, color='FFFFFF')
        cell.alignment = center_align
        cell.border = thin_border
    row += 1

    # ---- Row 3: summary data ----
    ws.cell(row=row, column=1, value=hierarchy_name).font = normal_font
    ws.cell(row=row, column=1).alignment = Alignment(horizontal='left', vertical='center')
    ws.cell(row=row, column=1).border = thin_border
    for c, key in enumerate(['LINE', 'TOGGLE', 'FSM', 'COND'], 2):
        cell = ws.cell(row=row, column=c)
        val = subtree.get(key, '--')
        if isinstance(val, (int, float)):
            cell.value = val
            cell.number_format = '0.0000'
        else:
            cell.value = val
        cell.font = normal_font
        cell.alignment = center_align
        cell.border = thin_border
    row += 1

    # ---- Sections ----
    sections = [
        ('Line Coverage',   ['Line', 'Code', 'Comment', 'Approve', 'Supplementary Materials'],
         [(ln, code) for ln, code in line_items]),
        ('Toggle Coverage', ['Type', 'Code', 'Comment', 'Approve', 'Supplementary Materials'],
         [(t, name) for t, name in toggle_items]),
        ('Cond Coverage',   ['Line', 'Code', 'Comment', 'Approve', 'Supplementary Materials'],
         [(ln, block) for ln, block in cond_items]),
        ('FSM Coverage',    ['Line', 'Transitions', 'Comment', 'Approve', 'Supplementary Materials'],
         [(ln, trans) for ln, trans in fsm_items]),
    ]

    for sec_title, sub_titles, items in sections:
        # Section header (merged)
        row = _write_section_header(row, sec_title)

        # Sub-header
        row = _write_sub_header(row, sub_titles)

        if not items:
            # Empty row with borders to show section is empty
            for c in range(1, 6):
                ws.cell(row=row, column=c).border = thin_border
            row += 1
        else:
            for item in items:
                ws.cell(row=row, column=1, value=item[0]).font = normal_font
                ws.cell(row=row, column=1).alignment = Alignment(horizontal='center', vertical='top')
                ws.cell(row=row, column=1).border = thin_border

                cell_b = ws.cell(row=row, column=2, value=item[1])
                cell_b.font = Font(name='Consolas', size=9)
                cell_b.alignment = wrap_align
                cell_b.border = thin_border

                for c in range(3, 6):
                    ws.cell(row=row, column=c).border = thin_border
                    ws.cell(row=row, column=c).font = normal_font
                row += 1

        # Blank separator row
        row += 1

    # ---- Column widths ----
    for c, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(c)].width = w

    # ---- Freeze panes ----
    ws.freeze_panes = 'A4'

    # ---- Save ----
    out_xlsx = os.path.join(OUTPUT_DIR, f"{target_module}.xlsx")
    wb.save(out_xlsx)
    log(f"  Excel report saved to {out_xlsx}")


# ---------------------------------------------------------------------------
# Step 3: split target module file by coverage type
# ---------------------------------------------------------------------------
def split_module_coverage(target_module: str) -> None:
    """
    Read output/module/<target_module>.txt and split it into per-coverage-type
    files under output/module_split/:

        <target_module>_line.txt
        <target_module>_cond.txt
        <target_module>_toggle.txt
        <target_module>_fsm.txt
    """
    module_file = os.path.join(MODULE_DIR, f"{target_module}.txt")
    if not os.path.exists(module_file):
        log(f"WARNING: Module file {module_file} not found — skipping split step")
        return

    log("=" * 60)
    log(f"Step 3: Splitting {target_module}.txt by coverage type")
    log("=" * 60)

    with open(module_file, 'r', encoding='utf-8', errors='replace') as fh:
        content = fh.read()

    lines = content.split('\n')
    SEP = '-' * 79

    # Map coverage type keyword -> output file suffix
    cov_type_map = {
        'Line':   'line',
        'Cond':   'cond',
        'Toggle': 'toggle',
        'FSM':    'fsm',
    }

    # Locate all separator lines (---...---) and the coverage type that follows
    # Section spans from the separator line to the next separator (or EOF).
    sep_positions = []          # list of (line_index, coverage_key)
    for i, line in enumerate(lines):
        if line == SEP and i + 1 < len(lines):
            # Check if next line is a coverage-type header
            match = re.match(r'^(\w+) Coverage for Module ', lines[i + 1])
            if match:
                cov_key = match.group(1)
                if cov_key in cov_type_map:
                    sep_positions.append((i, cov_key))

    # Extract each section
    extracted = {}
    for idx, (start, cov_key) in enumerate(sep_positions):
        end = sep_positions[idx + 1][0] if idx + 1 < len(sep_positions) else len(lines)
        section_text = '\n'.join(lines[start:end])
        extracted[cov_key] = section_text

        suffix = cov_type_map[cov_key]
        out_path = os.path.join(SPLIT_DIR, f"{target_module}_{suffix}.txt")
        with open(out_path, 'w', encoding='utf-8') as fh:
            fh.write(section_text)
        log(f"  Saved {out_path}  ({len(section_text)} chars)")

    # Report any missing types
    for cov_key in cov_type_map:
        if cov_key not in extracted:
            log(f"  WARNING: {cov_key} Coverage section not found in {target_module}.txt")


# ---------------------------------------------------------------------------
# main
# ---------------------------------------------------------------------------
def main() -> None:
    parser = argparse.ArgumentParser(description='Extract coverage data from VDB databases')
    parser.add_argument('-m', required=True, help='Module name to extract')
    args = parser.parse_args()
    target_module = args.m

    log(f"{'=' * 60}")
    log(f"Coverage Report Generation Started")
    log(f"  Target module : {target_module}")
    log(f"  Time          : {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    log(f"{'=' * 60}")

    # Step 1
    if not run_urg():
        log("ERROR: Script aborted — urg step failed")
        return

    # Step 2
    modules = parse_modinfo()
    if not modules:
        log("ERROR: Script aborted — no modules extracted from modinfo.txt")
        return

    # Step 3: split target module by coverage type
    split_module_coverage(target_module)

    # Step 4
    generate_excel_from_template(target_module)

    log("=" * 60)
    log("Coverage Report Generation Completed Successfully")
    log("=" * 60)


if __name__ == '__main__':
    main()
